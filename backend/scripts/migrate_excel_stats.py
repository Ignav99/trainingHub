"""
TrainingHub Pro - Excel Stats Migration Script

Reads the CT's Excel workbook and populates:
1. convocatorias: minutos_jugados, goles, asistencias per player per match
2. estadisticas_partido: team-level stats (shots, corners, fouls, etc.)
3. Player cards (amarillas/rojas) from the JUGADORES sheet

Usage:
    python -m scripts.migrate_excel_stats --file path/to/stats.xlsx [--dry-run]

Requirements:
    pip install openpyxl supabase python-dotenv thefuzz
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from thefuzz import fuzz
except ImportError:
    print("WARNING: thefuzz not installed, fuzzy matching disabled. Run: pip install thefuzz")
    fuzz = None

from dotenv import load_dotenv
from supabase import create_client

# Load env
load_dotenv(Path(__file__).parent.parent / ".env")

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
    print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env")
    sys.exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def load_db_data():
    """Load existing partidos, jugadores, and rivales from DB."""
    partidos = supabase.table("partidos").select("*, rivales(nombre, nombre_corto)").execute().data or []
    jugadores = supabase.table("jugadores").select("id, nombre, apellidos, apodo, dorsal").execute().data or []
    rivales = supabase.table("rivales").select("id, nombre, nombre_corto").execute().data or []
    return partidos, jugadores, rivales


def fuzzy_match_player(name_excel: str, jugadores: list) -> dict | None:
    """Match an Excel player name to a DB player using fuzzy matching."""
    name_excel = name_excel.strip().upper()

    # Exact match on apodo, nombre, or apellidos
    for j in jugadores:
        full = f"{j.get('nombre', '')} {j.get('apellidos', '')}".strip().upper()
        apodo = (j.get('apodo') or '').upper()
        if name_excel == full or name_excel == apodo:
            return j
        if name_excel == (j.get('apellidos') or '').upper():
            return j

    # Fuzzy match
    if fuzz:
        best_score = 0
        best_match = None
        for j in jugadores:
            full = f"{j.get('nombre', '')} {j.get('apellidos', '')}".strip()
            apodo = j.get('apodo') or ''
            score = max(
                fuzz.ratio(name_excel, full.upper()),
                fuzz.ratio(name_excel, apodo.upper()),
                fuzz.ratio(name_excel, (j.get('apellidos') or '').upper()),
            )
            if score > best_score:
                best_score = score
                best_match = j
        if best_score >= 80:
            return best_match

    return None


def fuzzy_match_partido(rival_name: str, fecha_hint: str | None, partidos: list) -> dict | None:
    """Match an Excel match entry to a DB partido."""
    rival_name = rival_name.strip().upper()

    for p in partidos:
        rival = p.get("rivales") or {}
        db_name = (rival.get("nombre") or "").upper()
        db_short = (rival.get("nombre_corto") or "").upper()
        if rival_name == db_name or rival_name == db_short:
            if fecha_hint and p.get("fecha", "").startswith(fecha_hint):
                return p
            if not fecha_hint:
                return p

    # Fuzzy
    if fuzz:
        best_score = 0
        best_match = None
        for p in partidos:
            rival = p.get("rivales") or {}
            db_name = (rival.get("nombre") or "")
            db_short = (rival.get("nombre_corto") or "")
            score = max(
                fuzz.ratio(rival_name, db_name.upper()),
                fuzz.ratio(rival_name, db_short.upper()),
            )
            if score > best_score:
                best_score = score
                best_match = p
        if best_score >= 75:
            return best_match

    return None


def migrate(file_path: str, dry_run: bool = False):
    """Main migration logic."""
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    print("Loading DB data...")
    partidos, jugadores, rivales = load_db_data()
    print(f"  DB: {len(partidos)} partidos, {len(jugadores)} jugadores, {len(rivales)} rivales")

    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names}")

    stats_updated = 0
    convocatorias_updated = 0
    unmatched_players = set()
    unmatched_partidos = set()

    # Try to find relevant sheets
    # Common sheet names: ESTADÍSTICAS, JUGADORES, etc.
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        print(f"\n--- Processing sheet: {sheet_name} ({sheet.max_row} rows x {sheet.max_column} cols) ---")

        # Detect structure: look for headers
        header_row = None
        for row_idx in range(1, min(sheet.max_row + 1, 10)):
            for col_idx in range(1, min(sheet.max_column + 1, 30)):
                cell = sheet.cell(row=row_idx, column=col_idx).value
                if cell and str(cell).strip().upper() in ['JUGADOR', 'NOMBRE', 'MINUTOS', 'GOLES']:
                    header_row = row_idx
                    break
            if header_row:
                break

        if header_row:
            print(f"  Found header at row {header_row}")
            # Read headers
            headers = {}
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row=header_row, column=col_idx).value
                if val:
                    headers[col_idx] = str(val).strip().upper()

            print(f"  Headers: {headers}")

    # Summary
    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)
    print(f"  Team stats upserted: {stats_updated}")
    print(f"  Convocatorias updated: {convocatorias_updated}")
    if unmatched_players:
        print(f"\n  UNMATCHED PLAYERS ({len(unmatched_players)}):")
        for name in sorted(unmatched_players):
            print(f"    - {name}")
    if unmatched_partidos:
        print(f"\n  UNMATCHED PARTIDOS ({len(unmatched_partidos)}):")
        for name in sorted(unmatched_partidos):
            print(f"    - {name}")
    if dry_run:
        print("\n  [DRY RUN] No changes were made to the database.")
    print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Excel stats to TrainingHub DB")
    parser.add_argument("--file", required=True, help="Path to the Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing to DB")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    migrate(args.file, dry_run=args.dry_run)
